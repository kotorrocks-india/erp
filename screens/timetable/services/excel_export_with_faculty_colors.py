# ============================================================================
# EXCEL EXPORT WITH FACULTY ROLE-BASED COLORS (YAML Compliant)
# ============================================================================
# Complete implementation of Slide 23 YAML spec for Excel export
# Features:
# - Faculty role coloring: Blue (In-Charge), Pink (Visiting), Black (Regular)
# - Extended afternoon marker (*)
# - All-day elective marker (†)
# - Proper cell merging for bridged subjects
# - Year-based row colors
# - Legend with color coding
# ============================================================================

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.engine import Engine
from typing import Dict, List, Optional
from datetime import datetime


# ============================================================================
# COLOR DEFINITIONS (from config.py YAML spec)
# ============================================================================

FACULTY_COLORS = {
    'in_charge': '2F80ED',      # Blue - Subject In-Charge (first faculty)
    'regular': '000000',         # Black - Regular Faculty  
    'visiting': 'FF6AA2',        # Pink - Visiting Faculty
}

BACKGROUND_COLORS = {
    'header': 'D3D3D3',          # Light gray for headers
    'year_1': 'FFE6E6',          # Light red/pink
    'year_2': 'E6F0FF',          # Light blue
    'year_3': 'E6FFE6',          # Light green
    'year_4': 'FFF4E6',          # Light orange
    'year_5': 'F0E6FF',          # Light purple
    'extended': 'FFF9C4',        # Light yellow for extended (*) periods
    'all_day': 'E3F2FD',        # Light blue for all-day (†) electives
}

MARKER_SYMBOLS = {
    'extended': '*',             # Extended afternoon (beyond P8)
    'all_day': '†',             # All-day elective block
}


# ============================================================================
# EXCEL EXPORT CLASS
# ============================================================================

class TimetableExcelExporter:
    """Export timetable to Excel with faculty role-based colors"""
    
    def __init__(self, engine: Engine):
        self.engine = engine
    
    def export_to_excel(
        self,
        context: Dict,
        filename: str = None
    ) -> str:
        """
        Export timetable to Excel file
        
        Args:
            context: Dict with ay_label, degree_code, year, term, division_code
            filename: Output filename (auto-generated if None)
            
        Returns:
            Path to created Excel file
        """
        
        # Auto-generate filename if not provided
        if not filename:
            filename = (
                f"TT_{context['degree_code']}_{context['ay_label']}_"
                f"Y{context['year']}_T{context['term']}"
            )
            if context.get('division_code'):
                filename += f"_{context['division_code']}"
            filename += f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Y{context['year']}_T{context['term']}"
        
        # Fetch data
        days_data = self._fetch_days_data(context)
        
        # Render each day
        current_row = 1
        
        for day_data in days_data:
            day_name = day_data['day_name']
            day_num = day_data['day_num']
            periods = day_data['periods']
            
            # Day header
            current_row = self._render_day_header(ws, current_row, day_name, len(periods))
            
            # Period headers
            current_row = self._render_period_headers(ws, current_row, periods)
            
            # Year rows (1-5 for BARCH)
            for year in range(1, 6):
                slots = self._fetch_slots_for_year(context, year, day_num)
                current_row = self._render_year_row(
                    ws, current_row, year, periods, slots, context
                )
            
            # Add spacing between days
            current_row += 1
        
        # Add legend
        self._add_legend(ws, current_row + 1)
        
        # Auto-size columns
        self._auto_size_columns(ws)
        
        # Save
        wb.save(filename)
        return filename
    
    def _fetch_days_data(self, context: Dict) -> List[Dict]:
        """Fetch all days and their periods"""
        
        with self.engine.connect() as conn:
            
            # Get day template for this context
            template_result = conn.execute(text("""
                SELECT id FROM day_templates
                WHERE ay_label = :ay
                  AND degree_code = :deg
                  AND term = :term
                  AND status = 'published'
                LIMIT 1
            """), {
                'ay': context['ay_label'],
                'deg': context['degree_code'],
                'term': context['term']
            }).fetchone()
            
            if not template_result:
                return []
            
            template_id = template_result[0]
            
            # Get all days (Mon-Sat)
            days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
            days_data = []
            
            for day_num, day_name in enumerate(days, 1):
                
                # Get periods for this day
                periods_result = conn.execute(text("""
                    SELECT 
                        slot_index,
                        kind_code,
                        label,
                        start_time,
                        end_time,
                        is_teaching_slot
                    FROM day_template_slots
                    WHERE template_id = :tid
                      AND is_teaching_slot = 1
                    ORDER BY slot_index
                """), {'tid': template_id}).fetchall()
                
                periods = [
                    {
                        'id': p[0],
                        'kind': p[1],
                        'label': p[2],
                        'start_time': p[3],
                        'end_time': p[4],
                        'is_teaching': p[5]
                    }
                    for p in periods_result
                ]
                
                days_data.append({
                    'day_num': day_num,
                    'day_name': day_name,
                    'periods': periods
                })
            
            return days_data
    
    def _fetch_slots_for_year(
        self, 
        context: Dict, 
        year: int, 
        day_num: int
    ) -> List[Dict]:
        """Fetch all slots for a year/day"""
        
        with self.engine.connect() as conn:
            
            result = conn.execute(text("""
                SELECT 
                    s.*,
                    -- Get faculty info
                    (SELECT faculty_in_charge FROM subject_offerings
                     WHERE id = s.offering_id) as subject_in_charge,
                    (SELECT faculty_list FROM subject_offerings
                     WHERE id = s.offering_id) as faculty_list_json,
                    -- Get faculty affiliations to determine visiting status
                    (SELECT affiliation_type FROM faculty_affiliations
                     WHERE faculty_email = (
                         SELECT faculty_in_charge FROM subject_offerings
                         WHERE id = s.offering_id
                     ) AND degree_code = :deg
                     LIMIT 1) as in_charge_affiliation
                     
                FROM timetable_slots s
                WHERE s.ay_label = :ay
                  AND s.degree_code = :deg
                  AND s.year = :year
                  AND s.term = :term
                  AND s.division_code = :div
                  AND s.day_of_week = :day
                  AND s.status = 'published'
                ORDER BY s.period_id
            """), {
                'ay': context['ay_label'],
                'deg': context['degree_code'],
                'year': year,
                'term': context['term'],
                'div': context.get('division_code', 'A'),
                'day': day_num
            }).fetchall()
            
            return [dict(row._mapping) for row in result]
    
    def _render_day_header(
        self, 
        ws, 
        row: int, 
        day_name: str, 
        num_periods: int
    ) -> int:
        """Render day header row"""
        
        cell = ws.cell(row=row, column=1, value=day_name)
        cell.font = Font(bold=True, size=14, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge across all period columns + year label column
        ws.merge_cells(
            start_row=row, 
            start_column=1, 
            end_row=row, 
            end_column=num_periods + 1
        )
        
        return row + 1
    
    def _render_period_headers(
        self, 
        ws, 
        row: int, 
        periods: List[Dict]
    ) -> int:
        """Render period header row"""
        
        # Year label column
        cell = ws.cell(row=row, column=1, value="Year / Period")
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(
            start_color=BACKGROUND_COLORS['header'], 
            end_color=BACKGROUND_COLORS['header'], 
            fill_type='solid'
        )
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Period columns
        for idx, period in enumerate(periods, 2):
            cell = ws.cell(
                row=row, 
                column=idx, 
                value=f"{period['label']}\n{period['start_time']}-{period['end_time']}"
            )
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(
                start_color=BACKGROUND_COLORS['header'], 
                end_color=BACKGROUND_COLORS['header'], 
                fill_type='solid'
            )
            cell.alignment = Alignment(
                horizontal='center', 
                vertical='center', 
                wrap_text=True
            )
        
        return row + 1
    
    def _render_year_row(
        self, 
        ws, 
        row: int, 
        year: int, 
        periods: List[Dict], 
        slots: List[Dict],
        context: Dict
    ) -> int:
        """
        Render one year row with faculty role-based coloring
        
        CRITICAL: Implements YAML color spec:
        - Blue (in_charge) for Subject In-Charge
        - Pink (visiting) for Visiting Faculty
        - Black (regular) for Regular Faculty
        """
        
        # Year label
        year_bg = BACKGROUND_COLORS.get(f'year_{year}', 'FFFFFF')
        semester = ['I', 'III', 'V', 'VII', 'IX'][year - 1]
        ordinal = ['1st', '2nd', '3rd', '4th', '5th'][year - 1]
        
        cell = ws.cell(row=row, column=1, value=f"{ordinal} Year Semester {semester}")
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(
            start_color=year_bg, 
            end_color=year_bg, 
            fill_type='solid'
        )
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Build period -> slot map (handle bridges)
        period_map = {}
        processed_bridges = set()
        
        for slot in slots:
            pid = slot['period_id']
            bridge_id = slot.get('bridge_group_id')
            
            if bridge_id:
                if bridge_id not in processed_bridges and slot['bridge_position'] == 1:
                    period_map[pid] = {
                        'type': 'bridge',
                        'slot': slot,
                        'colspan': slot['bridge_length']
                    }
                    processed_bridges.add(bridge_id)
                # Skip non-first bridge cells
            else:
                period_map[pid] = {
                    'type': 'single',
                    'slot': slot
                }
        
        # Render cells
        col = 2  # Start after year label
        
        for period in periods:
            pid = period['id']
            
            if pid in period_map:
                info = period_map[pid]
                slot = info['slot']
                
                # Determine cell content
                subject_name = slot.get('subject_name', slot['subject_code'])
                
                # Get faculty info
                in_charge_email = slot.get('subject_in_charge')
                faculty_list_json = slot.get('faculty_list_json')
                in_charge_affiliation = slot.get('in_charge_affiliation', 'core')
                
                # Parse faculty list
                import json
                try:
                    faculty_list = json.loads(faculty_list_json) if faculty_list_json else []
                except:
                    faculty_list = []
                
                # Determine if extended (period > 8)
                is_extended = pid > 8
                
                # Determine if all-day (check if slot has all_day_block flag)
                is_all_day = slot.get('is_all_day_block', False)
                
                # Build cell content with faculty names
                cell_content = subject_name
                
                # Add markers
                if is_extended:
                    cell_content += f" {MARKER_SYMBOLS['extended']}"
                if is_all_day:
                    cell_content += f" {MARKER_SYMBOLS['all_day']}"
                
                # Add faculty names with role-based coloring
                cell_content += "\n"
                
                # Get faculty with affiliations
                faculty_with_colors = self._get_faculty_with_colors(
                    in_charge_email,
                    faculty_list,
                    context
                )
                
                for fac_info in faculty_with_colors:
                    cell_content += f"\n{fac_info['name']}"
                
                # Create cell (handle colspan for bridges)
                if info['type'] == 'bridge':
                    # Merge cells for bridge
                    ws.merge_cells(
                        start_row=row,
                        start_column=col,
                        end_row=row,
                        end_column=col + info['colspan'] - 1
                    )
                
                cell = ws.cell(row=row, column=col, value=cell_content)
                
                # Set background color
                bg_color = year_bg
                if is_all_day:
                    bg_color = BACKGROUND_COLORS['all_day']
                elif is_extended:
                    bg_color = BACKGROUND_COLORS['extended']
                
                cell.fill = PatternFill(
                    start_color=bg_color,
                    end_color=bg_color,
                    fill_type='solid'
                )
                
                # CRITICAL: Apply faculty role-based text colors
                # This is done via rich text (requires openpyxl.cell.text.InlineFont)
                self._apply_faculty_colors_to_cell(
                    cell, 
                    subject_name, 
                    faculty_with_colors,
                    is_extended,
                    is_all_day
                )
                
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='top',
                    wrap_text=True
                )
                
                # Add border
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
                
                # Move column pointer
                if info['type'] == 'bridge':
                    col += info['colspan']
                else:
                    col += 1
            
            else:
                # Empty cell
                cell = ws.cell(row=row, column=col, value="")
                cell.fill = PatternFill(
                    start_color=year_bg,
                    end_color=year_bg,
                    fill_type='solid'
                )
                col += 1
        
        return row + 1
    
    def _get_faculty_with_colors(
        self,
        in_charge_email: str,
        faculty_list: List[str],
        context: Dict
    ) -> List[Dict]:
        """
        Get faculty with their role-based color codes
        
        Returns:
            List of {'name': str, 'color': str, 'role': str}
        """
        
        faculty_with_colors = []
        
        if not in_charge_email and not faculty_list:
            return []
        
        # Ensure in-charge is first
        all_faculty = [in_charge_email] if in_charge_email else []
        if faculty_list:
            all_faculty.extend([f for f in faculty_list if f != in_charge_email])
        
        # Get affiliations from database
        with self.engine.connect() as conn:
            
            for idx, email in enumerate(all_faculty):
                
                # Get affiliation
                result = conn.execute(text("""
                    SELECT affiliation_type, faculty_name
                    FROM faculty_affiliations
                    WHERE faculty_email = :email
                      AND degree_code = :deg
                    LIMIT 1
                """), {
                    'email': email,
                    'deg': context['degree_code']
                }).fetchone()
                
                if result:
                    affiliation = result[0]
                    name = result[1] or email.split('@')[0].replace('.', ' ').title()
                else:
                    affiliation = 'core'
                    name = email.split('@')[0].replace('.', ' ').title()
                
                # Determine color
                if idx == 0:
                    # First faculty = In-Charge (Blue)
                    color = FACULTY_COLORS['in_charge']
                    role = 'in_charge'
                elif affiliation == 'visiting':
                    # Visiting faculty (Pink)
                    color = FACULTY_COLORS['visiting']
                    role = 'visiting'
                else:
                    # Regular faculty (Black)
                    color = FACULTY_COLORS['regular']
                    role = 'regular'
                
                faculty_with_colors.append({
                    'name': name,
                    'color': color,
                    'role': role
                })
        
        return faculty_with_colors
    
    def _apply_faculty_colors_to_cell(
        self,
        cell,
        subject_name: str,
        faculty_with_colors: List[Dict],
        is_extended: bool,
        is_all_day: bool
    ):
        """
        Apply rich text formatting with faculty role-based colors
        
        CRITICAL: This implements the YAML color spec
        """
        
        from openpyxl.cell.text import InlineFont
        from openpyxl.cell.rich_text import TextBlock, CellRichText
        
        text_blocks = []
        
        # Subject name (black, bold)
        subject_text = subject_name
        if is_extended:
            subject_text += f" {MARKER_SYMBOLS['extended']}"
        if is_all_day:
            subject_text += f" {MARKER_SYMBOLS['all_day']}"
        
        text_blocks.append(
            TextBlock(
                InlineFont(b=True, sz=11, color='000000'),
                subject_text + "\n"
            )
        )
        
        # Faculty names with role-based colors
        for fac in faculty_with_colors:
            text_blocks.append(
                TextBlock(
                    InlineFont(
                        sz=10,
                        color=fac['color'],  # CRITICAL: Role-based color
                        b=(fac['role'] == 'in_charge')  # Bold for in-charge
                    ),
                    f"\n{fac['name']}"
                )
            )
        
        # Apply rich text
        cell.value = CellRichText(*text_blocks)
    
    def _add_legend(self, ws, row: int):
        """Add color legend at bottom"""
        
        cell = ws.cell(row=row, column=1, value="Legend:")
        cell.font = Font(bold=True, size=11)
        row += 1
        
        legends = [
            (f"● Subject In-Charge", FACULTY_COLORS['in_charge']),
            (f"● Regular Faculty", FACULTY_COLORS['regular']),
            (f"● Visiting Faculty", FACULTY_COLORS['visiting']),
            (f"{MARKER_SYMBOLS['extended']} Extended Afternoon", None),
            (f"{MARKER_SYMBOLS['all_day']} All-Day Elective", None),
        ]
        
        for text, color in legends:
            cell = ws.cell(row=row, column=1, value=text)
            if color:
                cell.font = Font(size=10, color=color, bold=True)
            else:
                cell.font = Font(size=10)
            row += 1
    
    def _auto_size_columns(self, ws):
        """Auto-size all columns"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width


# ============================================================================
# USAGE EXAMPLE
# ============================================================================

def export_timetable_to_excel(context: Dict, engine: Engine) -> str:
    """
    Export timetable with faculty role-based colors
    
    Args:
        context: Dict with ay_label, degree_code, year, term, division_code
        engine: Database engine
        
    Returns:
        Path to created Excel file
    """
    
    exporter = TimetableExcelExporter(engine)
    filepath = exporter.export_to_excel(context)
    
    return filepath


# Export for use in UI
__all__ = ['TimetableExcelExporter', 'export_timetable_to_excel', 'FACULTY_COLORS', 'BACKGROUND_COLORS']
